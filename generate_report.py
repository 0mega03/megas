#!/usr/bin/env python3
"""営業定例MTG資料を Notion の各DB（顧客・受注・見積）と月次PL（xlsx）から集計し、
指定した Notion ページに「全社着地と進捗・ファネル＋ヨミ精度・前回アクション・個別案件・原因分析」を生成する。

使い方:
  python3 generate_report.py [--current-month 2026-07] [--dry-run]

トークンは既定で .tmp/notion_token.txt の「アクセストークン=...」行から読む。
売上目標は 01.division/PL の月次PL（Monthly_Input タブ・売上小計行）から取得する（単位：万円）。
受注はDB_受注情報（確定＝受注日が該当月／見込＝受注日未入力を当月計上）。ご提案済みヨミはDB_見積情報管理（受注/close＝継続・着地予想月別）。
"""
import argparse, urllib.request, json, sys, os, time, glob
from collections import defaultdict, Counter
from datetime import date, datetime

# ===== プロジェクト固定設定（VJ 営業DB）=====
DB = {
    'customer': 'cd305332579a8291ae2301d1744dcb0e',
    'quote':    '24305332-579a-8256-b325-0178de5b552c',
    'order':    '2f005332-579a-83bf-8b35-01fe09c84aad',
    'goal':     '37b05332-579a-81ec-aac2-d6fff5706222',
}
# 出力先（MTG DB）。新規ページを作成し、そのページ本文に資料を書き込む
OUTPUT_DB = '24f05332-579a-8047-a00d-f092eb1c6c2c'
AGENDA_NAME_PREFIX = '営業定例アジェンダ-'  # + YYYYMMDD（作成日）
AGENDA_PLACE = 'VISIONARY JAPAN本社'        # 場所(select)
AGENDA_CATEGORY = '営業定例'                # カテゴリ(select)
# 担当の正は顧客DB「営業担当者」(people)。担当者名(formula)はその派生でフォールバック。
PERSON_MAP = {'妻鹿一大': '妻鹿', '古屋喬士': '喬士', 'Mega': '妻鹿', 'furuyatakashi': '喬士'}
MAIN = ['妻鹿', '喬士']                 # 個人別に表示する担当（事業部）
FISCAL_START = 3                        # 会計年度開始月（3月始まり）

# ===== 売上目標（月次PL xlsx）=====
# 既定パス（見つからなければ 01.division/PL/*月次PL*.xlsx の最新をフォールバック）。単位：万円。
PL_FILE = '01.division/PL/20260708-AI_Technology_事業部_月次PL（個人） のコピー.xlsx'
PL_SHEET = 'Monthly_Input'
PL_MONTH_HEADER_ROW = 3   # C3:N3 に各月（datetime）
PL_TARGET_ROW = 58        # 「売上 小計」行（月間目標）
PL_FIRST_COL = 3          # C
PL_LAST_COL = 14          # N

# 商談フェーズ（実DBスキーマ準拠 2026-07）。継続延長交渉中はファネル対象外（受注後の継続交渉）。
# ※旧スキーマの「提案済」「提案済（役職者合意）」「稟議中」は廃止済み。提案済は3区分。
PROP_ORDER = ['提案/見積り準備', '提案済（合意なし）', '提案済（担当者合意）', '提案済（決済者合意[口頭受注]）']
DOWNSTREAM = ['契約手続き（発注書未締結）', '受注（初回締結完了）']  # 提案以降の到達累計に含める
FIRST6 = ['初回商談済_お断り/保留', '初回商談済_先方社内確認', '初回商談済_実現性調査',
    '初回商談済_NDA締結+調査', '初回商談済_再訪調整中', '初回商談済_再訪設定済']
PHASE_ROWS = ['初回再調整', '商談化'] + FIRST6 + PROP_ORDER + DOWNSTREAM
# ファネル：1・2・3（初回商談済6フェーズ集約）はステージ化。4（提案）は最重要のため4-1〜4-4に分解し纏めない。
STAGES = [
    ('1 初回再調整', ['初回再調整']),
    ('2 商談化', ['商談化']),
    ('3 初回商談済', list(FIRST6)),
    ('4-1 提案/見積り準備', ['提案/見積り準備']),
    ('4-2 提案済（合意なし）', ['提案済（合意なし）']),
    ('4-3 提案済（担当者合意）', ['提案済（担当者合意）']),
    ('4-4 提案済（決済者合意[口頭受注]）', ['提案済（決済者合意[口頭受注]）']),
    ('5 契約手続き', ['契約手続き（発注書未締結）']),
    ('6 受注', ['受注（初回締結完了）']),
]
WON_PHASE = '受注（初回締結完了）'  # 見積・顧客DBの商談フェーズがこれ＝確定情報
# 受注確度（提案後ヨミ定義）。ヨミ期待値＝見積金額×確度。未設定は0（額面のみ計上）。
CONF_ROWS = ['A：確度80%', 'B：確度50%', 'C：確度30%']
CONF_W = {'A：確度80%': 0.8, 'B：確度50%': 0.5, 'C：確度30%': 0.3}
# 接触レベル（ヨミ精度・加重）。顧客DB「接触者」を A/B/C/D に読み替えて件数加重する。
CONTACT_LEVELS = [
    ('A：決裁者が前向き', 0.9),
    ('B：決裁者と接点あり', 0.6),
    ('C：キーマンと接点あり', 0.3),
    ('D：担当者どまり', 0.1),
]
CONTACT_MAP = {'決裁者': 'A：決裁者が前向き', '役職者': 'B：決裁者と接点あり',
    'キーマン': 'C：キーマンと接点あり', '担当者': 'D：担当者どまり'}
# 個別案件（停滞）の並び順・深度
DEPTH = {'契約手続き（発注書未締結）': 16, '提案済（決済者合意[口頭受注]）': 13,
    '提案済（担当者合意）': 11, '提案済（合意なし）': 10, '提案/見積り準備': 9, '初回商談済_NDA締結+調査': 7,
    '初回商談済_実現性調査': 6, '初回商談済_再訪調整中': 5, '初回商談済_再訪設定済': 5,
    '初回商談済_先方社内確認': 4, '商談化': 2, '初回再調整': 1}
# 提案以降（停滞案件の抽出対象）
PROPOSAL_PLUS = PROP_ORDER + ['契約手続き（発注書未締結）']
# 2-2b 展示会別移行率：母数がこの件数以上の流入経路を降順に出力（下限未満は省略しログ表示）
RYU_MIN_N = 5
# 展示会 開催日程（流入経路名 → (開始, 終了) ISO日付）。経過日数は開催最終日から起算。新規展示会は追記する。
RYU_EVENTS = {
    'NexTech Week2026春': ('2026-04-15', '2026-04-17'),
    'Eight EXPO 202606': ('2026-06-03', '2026-06-04'),
    'AI博覧会': ('2025-08-27', '2025-08-28'),
}
def event_suffix(name, ref=None):
    """流入経路名 → '（4/15~4/17）　※◯日経過'。開催日未登録なら空文字。経過は開催最終日→ref。"""
    ev = RYU_EVENTS.get(name)
    if not ev: return ''
    s = datetime.strptime(ev[0], '%Y-%m-%d').date(); e = datetime.strptime(ev[1], '%Y-%m-%d').date()
    days = ((ref or date.today()) - e).days
    md = lambda d: f"{d.month}/{d.day}"
    return f"（{md(s)}~{md(e)}）　※{days}日経過"

# ===== Notion API =====
TOKEN = None
def api(url, method='GET', data=None):
    r = urllib.request.Request(url, method=method)
    r.add_header('Authorization', 'Bearer ' + TOKEN)
    r.add_header('Notion-Version', '2022-06-28')
    r.add_header('Content-Type', 'application/json')
    if data is not None:
        r.data = json.dumps(data).encode()
    try:
        with urllib.request.urlopen(r) as x:
            return x.status, json.loads(x.read().decode())
    except urllib.error.HTTPError as e:
        return e.code, json.loads(e.read().decode())
def qall(did):
    out = []; cur = None
    while True:
        p = {'page_size': 100}
        if cur: p['start_cursor'] = cur
        s, b = api('https://api.notion.com/v1/databases/' + did + '/query', 'POST', p)
        if s != 200:
            sys.exit('DB取得失敗 %s: %s' % (did, b.get('message')))
        out += b['results']; cur = b.get('next_cursor')
        if not b.get('has_more'): break
    return out
def gc(bid):
    out = []; cur = None
    while True:
        u = 'https://api.notion.com/v1/blocks/' + bid + '/children?page_size=100'
        if cur: u += '&start_cursor=' + cur
        s, b = api(u)
        out += b['results']; cur = b.get('next_cursor')
        if not b.get('has_more'): break
    return out

# ===== helpers =====
def fstr(p): return (p.get('formula', {}) or {}).get('string')
def pname(p):
    """営業担当者(people) の先頭ユーザー名を返す（担当の正）。未設定は None"""
    us = p.get('people') or []
    return us[0].get('name') if us else None
def man(n): return f"{round((n or 0)/10000):,}万"     # 円 → 万
def yen2man(n): return round((n or 0) / 10000)         # 円 → 万（数値）
def tman(v): return f"{round(v):,}万"                  # 万（数値） → 表示
def pct(a, b): return '―' if b == 0 else f'{round(a/b*100,1)}%'
def who(tanto): return PERSON_MAP.get(tanto, 'その他')
def roll_sel(pr, name):
    """rollup プロパティ内の select 値を返す"""
    for it in pr.get(name, {}).get('rollup', {}).get('array', []):
        if it.get('type') == 'select' and it.get('select'):
            return it['select']['name']
    return None
def rollup_status(pr): return roll_sel(pr, '受注/close')

# ===== 会計月ユーティリティ =====
def fiscal_months(cur_ym):
    """会計年度(3月始まり)の 開始月〜当月+1 の 'YYYY-MM' リスト(昇順)を返す"""
    y, m = int(cur_ym[:4]), int(cur_ym[5:7])
    fy = y if m >= FISCAL_START else y - 1
    seq = [((FISCAL_START + i - 1) % 12) + 1 for i in range(12)]  # [3,4,..2]
    months = []
    yy = fy
    for mm in seq:
        months.append((yy if mm >= FISCAL_START else yy + 1, mm))
    res = [f'{yr:04d}-{mm:02d}' for (yr, mm) in months]
    end = f'{y:04d}-{m:02d}'
    idx = res.index(end) if end in res else len(res) - 1
    return res[:min(idx + 2, len(res))], fy
def fy_all_months(cur_ym):
    """当該会計年度の12ヶ月 'YYYY-MM'（昇順）"""
    fyy = int(cur_ym[:4]) if int(cur_ym[5:7]) >= FISCAL_START else int(cur_ym[:4]) - 1
    ms = []
    for i in range(12):
        mm = (FISCAL_START - 1 + i) % 12 + 1
        yy = fyy if mm >= FISCAL_START else fyy + 1
        ms.append(f'{yy:04d}-{mm:02d}')
    return ms, fyy
def quarter_of(ym):
    m = int(ym[5:7])
    pos = (m - FISCAL_START) % 12
    return pos // 3 + 1  # 1..4
def fy_range(cur_ym):
    fyy = int(cur_ym[:4]) if int(cur_ym[5:7]) >= FISCAL_START else int(cur_ym[:4]) - 1
    lo = f'{fyy:04d}-{FISCAL_START:02d}'
    hi = f'{fyy+1:04d}-{FISCAL_START-1:02d}' if FISCAL_START > 1 else f'{fyy:04d}-12'
    return lo, hi, fyy

# ===== 売上目標（xlsx）=====
def load_targets(path):
    """月次PL（Monthly_Input）から {YYYY-MM: 万} を返す。取得不可なら空 dict。"""
    try:
        import openpyxl
    except Exception:
        print('※ openpyxl 未導入のため売上目標は空欄で出力（pip install openpyxl）', file=sys.stderr)
        return {}
    if not path or not os.path.exists(path):
        cands = sorted(glob.glob('01.division/PL/*月次PL*.xlsx'))
        if not cands:
            print('※ 月次PLファイルが見つからず売上目標は空欄で出力', file=sys.stderr)
            return {}
        path = cands[-1]
    wb = openpyxl.load_workbook(path, data_only=True)
    if PL_SHEET not in wb.sheetnames:
        print(f'※ シート「{PL_SHEET}」が無く売上目標は空欄で出力', file=sys.stderr)
        return {}
    ws = wb[PL_SHEET]
    tgt = {}
    for col in range(PL_FIRST_COL, PL_LAST_COL + 1):
        hv = ws.cell(row=PL_MONTH_HEADER_ROW, column=col).value
        ym = hv.strftime('%Y-%m') if isinstance(hv, (datetime, date)) else None
        if not ym:
            continue
        v = ws.cell(row=PL_TARGET_ROW, column=col).value
        if isinstance(v, (int, float)):
            tgt[ym] = float(v)
    return tgt

# ===== 集計 =====
def collect(cur_ym, targets, rvc_exclude=True):
    custs = qall(DB['customer'])
    cust = []; cmap = {}
    for pg in custs:
        pr = pg['properties']
        naf = pr.get('NA日', {}).get('formula', {}) or {}
        o = {'tanto': pname(pr.get('営業担当者', {})) or fstr(pr.get('担当者名', {})),
             'phase': (pr.get('商談フェーズ', {}).get('select') or {}).get('name'),
             'level': (pr.get('接触者', {}).get('select') or {}).get('name'),
             'conf': (pr.get('提案後ヨミ定義', {}).get('select') or {}).get('name'),
             'st': (pr.get('受注/close/継続', {}).get('select') or {}).get('name'),
             'ryu': (pr.get('流入経路', {}).get('select') or {}).get('name'),
             'name': ''.join(t.get('plain_text', '') for t in (pr.get('顧客名', {}).get('title') or [])).strip(),
             'nad': naf.get('date', {}).get('start') if naf.get('type') == 'date' else naf.get('string'),
             'na': fstr(pr.get('NA', {})), 'id': pg['id']}
        cust.append(o); cmap[pg['id']] = o
    # 2章スコープ：事業部（MAIN＝妻鹿/喬士）のみ・RVC除外（着地集計と一貫）。組織＝MAIN合算
    scope = [c for c in cust if who(c['tanto']) in MAIN and not (rvc_exclude and c['ryu'] == 'RVC')]
    def rows(w):
        if w == '組織': return scope
        return [c for c in scope if who(c['tanto']) == w]

    # 受注/close/継続 ステータス分類：継続＝アクティブ、close＝close＋close(熱)
    def is_keiz(c): return c['st'] == '継続'
    def is_close(c): return c['st'] in ('close', 'close(熱)')
    # フェーズ現在件数（総数・継続・close）
    phase_cnt = {ph: {w: sum(c['phase'] == ph for c in rows(w)) for w in ['組織'] + MAIN} for ph in PHASE_ROWS}
    phase_keiz = {ph: {w: sum(c['phase'] == ph and is_keiz(c) for c in rows(w)) for w in ['組織'] + MAIN} for ph in PHASE_ROWS}
    phase_close = {ph: {w: sum(c['phase'] == ph and is_close(c) for c in rows(w)) for w in ['組織'] + MAIN} for ph in PHASE_ROWS}
    # 9ステージ到達累計（total・継続・close で並列に集計）
    def stage_cum(rs):
        here = [sum(r['phase'] in phs for r in rs) for _, phs in STAGES]
        cum = [0] * len(STAGES); acc = 0
        for i in range(len(STAGES) - 1, -1, -1):
            acc += here[i]; cum[i] = acc
        return cum
    def cum_set(rs):  # {'t':total,'k':継続,'c':close} の到達累計リスト
        return {'t': stage_cum(rs), 'k': stage_cum([c for c in rs if is_keiz(c)]),
                'c': stage_cum([c for c in rs if is_close(c)])}
    cums = {w: cum_set(rows(w)) for w in ['組織'] + MAIN}
    # 展示会（流入経路）別 到達累計（母数 RYU_MIN_N 以上を降順）
    ryu_cnt = Counter(c['ryu'] for c in scope if c['ryu'])
    ryus = [r for r, n in ryu_cnt.most_common() if n >= RYU_MIN_N]
    ryus_omitted = [(r, n) for r, n in ryu_cnt.most_common() if 0 < n < RYU_MIN_N]
    cums_ryu = {r: {w: cum_set([c for c in rows(w) if c['ryu'] == r]) for w in ['組織'] + MAIN} for r in ryus}
    # 接触レベル（A/B/C/D 加重・総数/継続/close）
    def lvl(pred):
        return {lbl: {w: sum(CONTACT_MAP.get(c['level']) == lbl and pred(c) for c in rows(w)) for w in ['組織'] + MAIN}
                for lbl, _ in CONTACT_LEVELS}
    level_cnt = lvl(lambda c: True); level_keiz = lvl(is_keiz); level_close = lvl(is_close)
    level_none = {w: sum(CONTACT_MAP.get(c['level']) is None for c in rows(w)) for w in ['組織'] + MAIN}

    # 受注DB → 個人別 受注累計・FY（§1-2）。RVC除外は表示側で実施。
    orders = qall(DB['order'])
    by_person = defaultdict(lambda: {'c': 0, 'a': 0})    # 累計（RVC込みの生値。表示側で除外）
    rvc = defaultdict(lambda: {'c': 0, 'a': 0})
    fy26 = defaultdict(lambda: {'c': 0, 'a': 0})
    fy_lo, fy_hi, _ = fy_range(cur_ym)
    for o in orders:
        pr = o['properties']; amt = pr.get('受注金額', {}).get('number') or 0
        d = (pr.get('受注日(契約締結日)', {}).get('date') or {}).get('start')
        rel = pr.get('DB_顧客情報', {}).get('relation') or []
        ci = cmap.get(rel[0]['id'], {}) if rel else {}
        w = who(ci.get('tanto'))
        by_person[w]['c'] += 1; by_person[w]['a'] += amt
        if ci.get('ryu') == 'RVC':
            rvc[w]['c'] += 1; rvc[w]['a'] += amt
        if d and fy_lo <= d[:7] <= fy_hi:
            fy26[w]['c'] += 1; fy26[w]['a'] += amt

    # 見積DB 単一ソース → 着地（§1-1）。商談フェーズ=受注（初回締結完了）＝確定、継続＝ヨミ（確度加重）。
    # 確定と継続は排他（着地予想月ベース）なので二重計上が起きない。MAIN・RVC除外。
    quotes = qall(DB['quote'])
    conf_month = defaultdict(lambda: {'c': 0, 'a': 0})                       # 確定 着地予想月ym別（§1-1）
    ym_yomi = defaultdict(lambda: {'c': 0, 'a': 0, 'e': 0.0})                # 継続 着地予想月ym別（額面a・期待値e）
    yomi_person = defaultdict(lambda: {'c': 0, 'a': 0, 'e': 0.0})            # 個人別 継続ヨミ合計（§1-2）
    yomi_items = defaultdict(list)                                          # [w] -> 案件（§4）
    cust_qamt = defaultdict(float)                                          # 顧客id -> 継続見積金額合計（§4 ヨミ案件）
    cust_qland = {}                                                         # 顧客id -> 着地予想ym（最も早い実月）
    for q in quotes:
        pr = q['properties']; amt = pr.get('見積金額', {}).get('number') or 0
        land = (pr.get('目標受注日', {}).get('date') or {}).get('start')  # 実スキーマは目標受注日（旧名『着地予想月(月末を記載)』は存在しない）
        rel = pr.get('DB_顧客情報', {}).get('relation') or []
        cid = rel[0]['id'] if rel else None
        ci = cmap.get(cid, {}) if cid else {}
        w = who(ci.get('tanto'))
        if w not in MAIN:
            continue
        if rvc_exclude and ci.get('ryu') == 'RVC':
            continue
        m = land[:7] if land else '未設定'
        ph = roll_sel(pr, '商談フェーズ'); st = rollup_status(pr)
        if ph == WON_PHASE:                                    # 確定（受注済・確度100%）
            conf_month[m]['c'] += 1; conf_month[m]['a'] += amt
        elif st == '継続':                                     # ヨミ（見込・確度加重）
            conf = ci.get('conf'); e = amt * CONF_W.get(conf, 0.0)
            ym_yomi[m]['c'] += 1; ym_yomi[m]['a'] += amt; ym_yomi[m]['e'] += e
            yomi_person[w]['c'] += 1; yomi_person[w]['a'] += amt; yomi_person[w]['e'] += e
            yomi_items[w].append({'name': ci.get('name'), 'amt': amt, 'land': m,
                'phase': ph, 'conf': conf, 'e': e})
            if cid:                                            # 顧客別に見積金額を合算（複数見積の重複を解消）
                cust_qamt[cid] += amt
                if m != '未設定' and (cid not in cust_qland or cust_qland[cid] == '未設定' or m < cust_qland[cid]):
                    cust_qland[cid] = m
                cust_qland.setdefault(cid, '未設定')

    # ヨミ案件（§4）：提案以降で継続の案件＋継続見積のある案件を統合。金額は見積情報管理から取得。close/close(熱)は除外。
    yomi_deals = {}
    for w in MAIN:
        cand = {c['id'] for c in rows(w) if c['phase'] in PROPOSAL_PLUS and c['st'] not in ('close', 'close(熱)')}
        cand |= {cid for cid in cust_qamt if who(cmap.get(cid, {}).get('tanto')) == w}
        deals = []
        for cid in cand:
            ci = cmap.get(cid, {})
            deals.append({'name': ci.get('name'), 'amt': cust_qamt.get(cid, 0),
                'land': cust_qland.get(cid, '未設定'), 'phase': ci.get('phase'),
                'nad': ci.get('nad'), 'na': ci.get('na')})
        deals.sort(key=lambda d: -(d['amt'] or 0))
        yomi_deals[w] = deals

    return dict(rows=rows, phase_cnt=phase_cnt, phase_keiz=phase_keiz, phase_close=phase_close,
        cums=cums, level_cnt=level_cnt, level_keiz=level_keiz, level_close=level_close, level_none=level_none,
        by_person=by_person, rvc=rvc, fy26=fy26, conf_month=conf_month,
        ym_yomi=ym_yomi, yomi_person=yomi_person, yomi_items=yomi_items, yomi_deals=yomi_deals,
        ryus=ryus, ryus_omitted=ryus_omitted, cums_ryu=cums_ryu,
        cust=cust, targets=targets, rvc_exclude=rvc_exclude)

# ===== ブロック生成 =====
def rt(s, b=False): return [{"type": "text", "text": {"content": str(s)}, "annotations": {"bold": b}}]
def h1(s): return {"object": "block", "type": "heading_1", "heading_1": {"rich_text": rt(s, True)}}
def h2(s): return {"object": "block", "type": "heading_2", "heading_2": {"rich_text": rt(s, True)}}
def h3(s): return {"object": "block", "type": "heading_3", "heading_3": {"rich_text": rt(s, True)}}
def para(s): return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": rt(s)}}
def bp(s): return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": rt(s, True)}}
def bullet(s): return {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": rt(s)}}
def callout(s, e, c): return {"object": "block", "type": "callout", "callout": {"rich_text": rt(s), "icon": {"type": "emoji", "emoji": e}, "color": c}}
def div(): return {"object": "block", "type": "divider", "divider": {}}
def table(h, rows):
    tr = [{"object": "block", "type": "table_row", "table_row": {"cells": [rt(x) for x in h]}}]
    for r in rows:
        tr.append({"object": "block", "type": "table_row", "table_row": {"cells": [rt(x) for x in r]}})
    return {"object": "block", "type": "table", "table": {"table_width": len(h), "has_column_header": True, "has_row_header": False, "children": tr}}

QLABEL = {1: "第1Q（3-5月）", 2: "第2Q（6-8月）", 3: "第3Q（9-11月）", 4: "第4Q（12-2月）"}

def stage_move_table(cmap):
    """ステージ 到達/継続/close/直前移行率/累計移行率 テーブル（組織/妻鹿/喬士）。cmap[w]＝{'t','k','c'} 到達累計リスト。
    直前＝到達÷ひとつ手前ステージの到達。累計＝到達÷ステージ1到達（初回商談からの通算）。"""
    rows_ = []
    for i, (nm, _) in enumerate(STAGES):
        row = [nm]
        for w in ['組織'] + MAIN:
            cm = cmap[w]
            prev = '―' if i == 0 else pct(cm['t'][i], cm['t'][i-1])
            cumr = '―' if i == 0 else pct(cm['t'][i], cm['t'][0])
            row += [f"{cm['k'][i]}/{cm['t'][i]}", str(cm['c'][i]), prev, cumr]
        rows_.append(row)
    trow = [f"通算（1→{len(STAGES)}）"]
    for w in ['組織'] + MAIN:
        cm = cmap[w]; trow += [f"{cm['k'][-1]}/{cm['t'][-1]}", str(cm['c'][-1]), "―", pct(cm['t'][-1], cm['t'][0])]
    rows_.append(trow)
    hdr = ["ステージ"]
    for w in ['組織'] + MAIN: hdr += [f"継続/到達({w})", f"close({w})", f"直前({w})", f"累計({w})"]
    return table(hdr, rows_)

def build_blocks(D, cur_ym):
    B = []
    targets = D['targets']
    cm = D['conf_month']; ymy = D['ym_yomi']
    fy_lo, fy_hi, fyy = fy_range(cur_ym)
    months, _ = fiscal_months(cur_ym)
    fy_all, _ = fy_all_months(cur_ym)
    def in_fy(m): return m != '未設定' and fy_lo <= m <= fy_hi
    def gross(d): return sum(ymy.get(x, {'a': 0})['a'] for x in d)
    def exp(d): return sum(ymy.get(x, {'e': 0})['e'] for x in d)

    # 当月の各値（確定＝見積phase受注、ヨミ＝継続。見積DB単一ソース・二重計上なし）
    cur_conf = cm.get(cur_ym, {'a': 0, 'c': 0})['a']
    cur_gross = ymy.get(cur_ym, {'a': 0})['a']; cur_exp = ymy.get(cur_ym, {'e': 0})['e']
    cur_land = cur_conf + cur_exp

    # --- 1. 全社着地と進捗 ---
    B += [h1("1. 全社着地と進捗")]
    B += [callout(
        f"当月{int(cur_ym[5:7])}月の着地見込は{man(cur_land)}（確定{man(cur_conf)}＋ヨミ期待値{man(cur_exp)}）。"
        f"ヨミ額面は{man(cur_gross)}。担当は{'/'.join(MAIN)}・流入経路RVC除外。", "🎯", "blue_background")]

    # 1-1 月次・四半期 着地（目標・GAP・組織）
    B += [h3("1-1. 月次・四半期 着地（目標・GAP・組織）")]
    hdr = ["期間", "確定", "ヨミ額面", "ヨミ期待値", "着地見込", "目標", "GAP"]
    trows = []
    def row_for(label, ms, tgt):
        conf = sum(cm.get(x, {'a': 0})['a'] for x in ms)
        g = gross(ms); e = exp(ms); land = conf + e
        land_m = yen2man(land)
        return [label,
            man(conf) if conf else "―",
            man(g) if g else "―",
            man(e) if e else "―",
            man(land) if land else "―",
            tman(tgt) if tgt else "―（PL未集約）",
            tman(tgt - land_m) if tgt else "―"]
    # 年間
    fy_ms = [m for m in fy_all]
    trows.append(row_for("FY26 年間", fy_ms, sum(t for m, t in targets.items() if in_fy(m))))
    # 四半期（表示範囲に月がある四半期のみ）
    for qn in sorted({quarter_of(m) for m in months}):
        qms = [m for m in fy_all if quarter_of(m) == qn]
        trows.append(row_for(QLABEL[qn], qms, sum(targets.get(m, 0) for m in qms)))
    # 月別（降順）
    for m in sorted(months, reverse=True):
        trows.append(row_for(f"{int(m[5:7])}月", [m], targets.get(m, 0)))
    B += [table(hdr, trows)]
    B += [para("DB_見積情報管理が単一ソース（着地予想月ベース）。確定＝商談フェーズ「受注（初回締結完了）」。ヨミ＝受注/close＝継続。"
        "ヨミ期待値＝見積金額×受注確度（提案後ヨミ定義 A80/B50/C30・未設定は0）。着地見込＝確定＋ヨミ期待値。確定とヨミは排他のため二重計上なし。"
        "目標は月次PL（Monthly_Input・売上小計・万円）。GAP＝目標−着地見込。RVC除外。")]

    # 1-2 実績（個人別）
    B += [h3("1-2. 実績（個人別）")]
    B += [para("目標は事業部単位で設定。実績は流入経路RVC（外部パートナー経由）を"
        + ("除外して集計。" if D['rvc_exclude'] else "含めて集計。"))]
    p_rows = []; org_ord = 0
    for w in MAIN + ['その他']:
        bpv = D['by_person'][w]; rv = D['rvc'][w]; fyv = D['fy26'][w]
        a = bpv['a'] - (rv['a'] if D['rvc_exclude'] else 0); c = bpv['c'] - (rv['c'] if D['rvc_exclude'] else 0)
        org_ord += a
        if w in MAIN:
            yp = D['yomi_person'][w]
            p_rows.append([w, f"{man(a)} / {c}件", f"{man(fyv['a'])} / {fyv['c']}件", f"{man(yp['a'])} / {yp['c']}件"])
        else:
            p_rows.append([w + "（社長案件等）", f"{man(a)} / {c}件", "―", "―"])
    B += [table(["担当", "受注（累計/件）", "うちFY26内", "継続ヨミ（見積/件）"], p_rows)]
    rvc_org = sum(D['rvc'][w]['a'] for w in D['rvc']); rvc_orgc = sum(D['rvc'][w]['c'] for w in D['rvc'])
    if D['rvc_exclude'] and rvc_orgc:
        B += [bullet(f"RVC除外後の組織 受注累計は{man(org_ord)}（RVC流入{rvc_orgc}件・{man(rvc_org)}を除外）。")]
    B += [div()]

    # --- 2. ファネル＋ヨミ精度 ---
    cums = D['cums']; phase_cnt = D['phase_cnt']
    B += [h1("2. ファネル＋ヨミ精度")]
    # 2-1 フェーズ別 件数（総数／継続／close） ＋ 移行率
    B += [h3("2-1. フェーズ別 件数（組織 / 妻鹿 / 喬士・継続/close内訳）")]
    phase_keiz = D['phase_keiz']; phase_close = D['phase_close']
    stage_idx = {}
    for nm, phs in STAGES:
        for ph in phs: stage_idx[ph] = nm.split()[0]
    crows = []
    for ph in PHASE_ROWS:
        row = [f"{stage_idx.get(ph,'')} {ph}"]
        for w in ['組織'] + MAIN:
            row += [f"{phase_keiz[ph][w]}/{phase_cnt[ph][w]}", str(phase_close[ph][w])]
        crows.append(row)
    tot_row = ["合計"]
    for w in ['組織'] + MAIN:
        tot_row += [f"{sum(phase_keiz[ph][w] for ph in PHASE_ROWS)}/{len(D['rows'](w))}",
            str(sum(phase_close[ph][w] for ph in PHASE_ROWS))]
    crows.append(tot_row)
    hdr_c = ["商談フェーズ"]
    for w in ['組織'] + MAIN: hdr_c += [f"継続({w})/{w}", f"close({w})"]
    B += [table(hdr_c, crows)]
    B += [para("各担当ごとに 継続（アクティブ）/総数 ・ close（close＋close(熱)）。継続＝進行中、close＝クローズ/保留。受注済は「6 受注」フェーズに計上。")]

    B += [bp("ファネル移行率（提案レンジ4分解・到達累計ベース／到達・継続・close併記）")]
    B += [stage_move_table(cums)]
    m0, m1 = MAIN[0], MAIN[1]
    r0 = pct(cums[m0]['t'][3], cums[m0]['t'][2]); r1 = pct(cums[m1]['t'][3], cums[m1]['t'][2]); rorg = pct(cums['組織']['t'][3], cums['組織']['t'][2])
    B += [callout(
        f"「3 初回商談済 → 4-1 提案」到達率は {m0}{r0} / {m1}{r1}（組織{rorg}）。提案化の個人差が最大の論点。",
        "🔎", "yellow_background")]
    B += [para("各セル「継続/到達」＝そのステージ以降にいる案件の累計（継続＝うちアクティブ）。close列はそのステージ以降のクローズ累計。"
        "直前＝到達÷ひとつ手前ステージの到達（1段の移行率）。累計＝到達÷ステージ1到達（初回再調整からの通算移行率）。"
        "ステージ3は初回商談済の6フェーズを集約。ステージ4（提案）は最重要のため 4-1 提案準備／4-2 合意なし／4-3 担当者合意／4-4 決済者合意[口頭受注] に分解し纏めない。"
        "時系列の移行履歴ではなくスナップショット近似。")]

    # 2-2 全体歩留り（リード→受注）＋ アクティブ（継続）/close 内訳
    B += [h3("2-2. 全体歩留り（リード→受注）")]
    phase_keiz = D['phase_keiz']; phase_close = D['phase_close']
    yrow = ["歩留り"]; krow = ["継続（アクティブ）"]; crow = ["close（＋熱）"]
    for w in ['組織'] + MAIN:
        won = phase_cnt[WON_PHASE][w]; tot = len(D['rows'](w))
        keiz = sum(phase_keiz[ph][w] for ph in PHASE_ROWS)
        clo = sum(phase_close[ph][w] for ph in PHASE_ROWS)
        yrow.append(f"{pct(won, tot)}（{won}/{tot}）")
        krow.append(f"{keiz}（{pct(keiz, tot)}）"); crow.append(f"{clo}（{pct(clo, tot)}）")
    B += [table(["区分", "組織"] + MAIN, [yrow, krow, crow])]
    B += [para("受注＝「受注（初回締結完了）」到達件数。母数＝全顧客件数。継続＝進行中アクティブ、close＝クローズ/保留（close＋close(熱)）。"
        "提案レンジの内訳は 2-1 のファネル移行率（4-1〜4-4）に一本化。")]

    # 2-2b 展示会（流入経路）別 ステージ移行率
    B += [h3("2-2b. 展示会別 ステージ移行率")]
    B += [para(f"流入経路（展示会）ごとに、2-1と同じファネル（提案レンジ4分解）で到達・移行率を出力（母数{RYU_MIN_N}件以上を降順）。"
        "見出し（M/D~M/D）は開催日程、経過日数は開催最終日から本日まで。母数が小さい経路は移行率のブレに注意。")]
    for r in D['ryus']:
        n = D['cums_ryu'][r]['組織']['t'][0]
        B += [bp(f"{r}{event_suffix(r)}（母数{n}件）")]
        B += [stage_move_table(D['cums_ryu'][r])]
    if D['ryus_omitted']:
        om = "、".join(f"{r}{n}件" for r, n in D['ryus_omitted'])
        B += [para(f"※ 母数{RYU_MIN_N}件未満のため省略: {om}")]

    # 2-3 ヨミ加重（接触レベル A/B/C/D）
    B += [h3("2-3. ヨミ加重（接触レベル A/B/C/D・継続/close内訳）")]
    lk = D['level_keiz']; lc = D['level_close']
    lrows = []
    for lbl, _ in CONTACT_LEVELS:
        row = [lbl]
        for w in ['組織'] + MAIN:
            row += [f"{lk[lbl][w]}/{D['level_cnt'][lbl][w]}", str(lc[lbl][w])]
        lrows.append(row)
    lrows.append(["未設定"] + [x for w in ['組織'] + MAIN for x in [f"―/{D['level_none'][w]}", "―"]])
    hdr_l = ["接触レベル"]
    for w in ['組織'] + MAIN: hdr_l += [f"継続({w})/{w}", f"close({w})"]
    B += [table(hdr_l, lrows)]
    def weighted(w): return sum(D['level_cnt'][lbl][w] * wt for lbl, wt in CONTACT_LEVELS)
    wts = "/".join(f"{lbl[0]}{wt}" for lbl, wt in CONTACT_LEVELS)
    B += [callout(
        f"件数加重（{wts}）は 組織{weighted('組織'):.1f}・{m0}{weighted(m0):.1f}・{m1}{weighted(m1):.1f}。"
        f"決裁者接点(A+B)の薄さが受注率に直結。", "💡", "gray_background")]
    B += [para("接触レベルは顧客DB「接触者」を A/B/C/D に読み替え。未入力は「未設定」。加重は件数×係数。")]
    B += [div()]

    # --- 3. 前回アクション棚卸し ---
    B += [h1("3. 前回アクション棚卸し")]
    B += [para("一次情報（営業部MTG議事録）が未取得のため、本セクションは空欄。議事録が取得でき次第、アクションリストを記載する。")]
    B += [table(["アクション", "担当", "期限（日付）", "状態"], [["", "", "", ""]])]
    B += [div()]

    # --- 4. 個別案件 深掘り（ヨミ案件） ---
    B += [h1("4. 個別案件 深掘り（ヨミ案件）")]
    for w in MAIN:
        B += [h2(w)]
        deals = D['yomi_deals'][w]
        if deals:
            B += [bp("ヨミ案件")]
            drows = []
            for d in deals[:12]:
                land = f"{int(d['land'][5:7])}月" if d['land'] != '未設定' else "未設定"
                nad = d['nad'][:10] if d['nad'] else "―"
                amt = man(d['amt']) if d['amt'] else "―（未見積）"
                drows.append([d['name'] or '―', amt, land, d['phase'] or '―', "", nad])
            B += [table(["顧客名", "金額", "着地予想", "商談フェーズ", "滞留日数", "次アクション（NA日）"], drows)]
            B += [para("金額は DB_見積情報管理の継続見積合計（複数見積は合算）。提案以降で継続の案件＋継続見積のある案件を統合（close/close(熱)は除外）。滞留日数は取得元未定のため現状は空欄。")]
        else:
            B += [para("該当案件なし。")]
    B += [div()]

    # --- 5. 達成に向けた原因とアクションの切り分け ---
    B += [h1("5. 達成に向けた原因とアクションの切り分け")]
    B += [para("問題は定量で分解（自動）。原因（真因）・課題・打ち手は営業判断で仕上げる。")]
    for w in MAIN:
        cum = cums[w]['t']
        prop_rate = pct(cum[3], cum[2])                    # 初回商談済→提案 到達率
        n_shanai = phase_cnt['初回商談済_先方社内確認'][w]  # 預かり滞留
        n_teian = phase_cnt['提案/見積り準備'][w]
        problem = (f"提案/見積り準備が{n_teian}件、先方社内確認が{n_shanai}件と滞留。"
            f"初回商談済→提案の到達率{prop_rate}。受注に必要な提案化が弱い。")
        B += [h2(w)]
        B += [table(["項目", "内容"],
            [["問題", problem], ["原因", ""], ["課題", ""], ["打ち手", ""]])]
    return B

# ===== ページ書き込み =====
def append_blocks(page_id, blocks, start=0):
    for i in range(start, len(blocks), 12):
        s, b = api('https://api.notion.com/v1/blocks/' + page_id + '/children', 'PATCH', {'children': blocks[i:i+12]})
        if s != 200:
            sys.exit('append失敗: ' + json.dumps(b, ensure_ascii=False)[:400])
        time.sleep(0.35)

def rewrite_page(page_id, blocks):
    """既存ページの本文を全削除して再構築（--page-id 指定時）。child_page・transcription は保持"""
    old = gc(page_id)
    for blk in old:
        if blk['type'] in ('child_page', 'transcription'):
            continue
        api('https://api.notion.com/v1/blocks/' + blk['id'], 'DELETE'); time.sleep(0.12)
    append_blocks(page_id, blocks)

def create_in_db(db_id, blocks, today, title_prefix=''):
    """MTG DB に新規ページを作成し、本文に資料を書き込む。作成したページIDを返す"""
    ymd = today.strftime('%Y%m%d')
    props = {
        '名前':     {'title': [{'text': {'content': title_prefix + AGENDA_NAME_PREFIX + ymd}}]},
        '日付':     {'date': {'start': today.strftime('%Y-%m-%d')}},
        '場所':     {'select': {'name': AGENDA_PLACE}},
        'カテゴリ': {'select': {'name': AGENDA_CATEGORY}},
    }
    s, b = api('https://api.notion.com/v1/pages', 'POST',
               {'parent': {'database_id': db_id}, 'properties': props, 'children': blocks[:20]})
    if s != 200:
        sys.exit('ページ作成失敗: ' + json.dumps(b, ensure_ascii=False)[:400])
    pid = b['id']
    append_blocks(pid, blocks, start=20)
    return pid, b.get('url')

def load_token(path):
    with open(path, encoding='utf-8') as f:
        for line in f:
            if 'アクセストークン' in line and '=' in line:
                return line.split('=', 1)[1].strip()
            if line.strip().startswith(('ntn_', 'secret_')):
                return line.strip()
    sys.exit('トークンが見つからない: ' + path)

def main():
    global TOKEN
    ap = argparse.ArgumentParser()
    ap.add_argument('--db-id', default=OUTPUT_DB, help='新規ページ作成先DB（既定=MTG DB）')
    ap.add_argument('--page-id', default=None, help='既存ページを上書きする場合に指定（DB作成の代わり）')
    ap.add_argument('--token-file', default='.tmp/notion_token.txt')
    ap.add_argument('--pl-file', default=PL_FILE, help='売上目標を読む月次PL xlsx（Monthly_Input・売上小計行）')
    ap.add_argument('--current-month', default=None, help='YYYY-MM（既定=今日）。集計の対象月')
    ap.add_argument('--no-rvc-exclude', action='store_true', help='RVC除外をしない')
    ap.add_argument('--title-prefix', default='', help='新規ページのタイトル接頭辞（例: 【テスト】）')
    ap.add_argument('--dry-run', action='store_true', help='集計結果を表示しページは作成/更新しない')
    a = ap.parse_args()
    TOKEN = load_token(a.token_file)
    cur_ym = a.current_month or datetime.now().strftime('%Y-%m')
    targets = load_targets(a.pl_file)
    D = collect(cur_ym, targets, rvc_exclude=not a.no_rvc_exclude)
    blocks = build_blocks(D, cur_ym)
    tgt_cur = targets.get(cur_ym)
    print(f'集計完了: 顧客{len(D["cust"])}件 / 生成ブロック{len(blocks)} / 当月{cur_ym} / '
          f'当月目標{tman(tgt_cur) if tgt_cur else "未取得"} / 目標月数{len(targets)}')
    if a.dry_run:
        print('--dry-run: ページ未作成')
        return
    if a.page_id:
        rewrite_page(a.page_id.replace('-', ''), blocks)
        print('既存ページ更新完了:', a.page_id)
    else:
        pid, url = create_in_db(a.db_id.replace('-', ''), blocks, datetime.now(), a.title_prefix)
        print('新規ページ作成完了:', pid)
        if url:
            print('URL:', url)

if __name__ == '__main__':
    main()
